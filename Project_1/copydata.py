from openpyxl import load_workbook, Workbook
from copy import copy
import os
from datetime import datetime
# Source file
source_file = r"E:\python\Project_1\HR_SHEET\output.xlsx"

# Destination file
destination_file = r"E:\python\HR_Backup_file\Daily_HR_Sheet.xlsx"

# Load source workbook
source_wb = load_workbook(source_file)
source_ws = source_wb["Sheet3"]

# If destination exists
if os.path.exists(destination_file):

    dest_wb = load_workbook(destination_file)

    # Create sheet if not exists
    if "Sheet1" not in dest_wb.sheetnames:
        dest_ws = dest_wb.create_sheet("Sheet1")
    else:
        dest_ws = dest_wb["Sheet1"]

else:
    # Create new workbook
    dest_wb = Workbook()
    dest_ws = dest_wb.active
    dest_ws.title = "Sheet1"

# Find next empty row
start_row = dest_ws.max_row + 1

# If sheet empty
if dest_ws.max_row == 1 and dest_ws["A1"].value is None:
    start_row = 1

# Get max column from source
max_col = source_ws.max_column

# Add COPY_DATE header only first time
if start_row == 1:
    dest_ws.cell(row=1, column=max_col + 1, value="COPY_DATE")

# Copy data with formatting
for row in source_ws.iter_rows():

    # Skip header while appending next time
    if start_row > 1 and row[0].row == 1:
        continue

    new_row = start_row + row[0].row - 1

    if start_row > 1:
        new_row -= 1

    for cell in row:

        new_cell = dest_ws.cell(
            row=new_row,
            column=cell.column,
            value=cell.value
        )

        # Copy formatting
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)

    # Add current date column
    date_cell = dest_ws.cell(
        row=new_row,
        column=max_col + 1,
        value=datetime.today().strftime("%Y-%m-%d")
    )

# Save file
dest_wb.save(destination_file)

#print("Data with formatting and COPY_DATE appended successfully.")