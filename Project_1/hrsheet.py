import pandas as pd
import sys
sys.path.append(r"E:\python")
from DbConfig.db_connectionlive import get_connection
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from excelsheetofmail import file_path
import warnings
warnings.filterwarnings("ignore")


# DB Connection
conn = get_connection()
if not conn:
    print("DB Connection Failed")
    exit()

# Step 1: Read Excel
#file_path = r"C:\Users\software.support\Desktop\HR_SHEET\input.xlsx"
df = pd.read_excel(file_path, sheet_name="Sheet1")

# Step 2: Prepare Employee Code list (clean)
codes = df['Employee Code'].dropna().astype(str).str.strip()

# Convert to SQL IN format → '1001','1002','1003'
sql_list = ",".join([f"'{c}'" for c in codes])

# Step 3: Create Query
query = f""" SELECT e.empno, e.emp_status, e.inactivedatetime, e.updatedby,
case when e.emp_status=76 then u.username else '' end  as SYSTEM_USER, e.inactivedatetime as SYSTEM_DATE
FROM employee e
left join hisuser u on u.id=e.updatedby
WHERE e.empno in ({sql_list}) """

#print("Executing Query...")

# Step 4: Execute and fetch result
result_df = pd.read_sql(query, conn)

# Convert both columns to string for matching
df['Employee Code'] = df['Employee Code'].astype(str).str.strip()
result_df['EMPNO'] = result_df['EMPNO'].astype(str).str.strip()
df.rename(columns={'Final Approved LWD': 'Final_Approved_LWD'}, inplace=True)
# Take required columns from SQL result
lookup_df = result_df[['EMPNO', 'EMP_STATUS', 'INACTIVEDATETIME','SYSTEM_USER','SYSTEM_DATE']]

# Merge (VLOOKUP equivalent)
merged_df = df.merge(
    lookup_df,
    left_on='Employee Code',
    right_on='EMPNO',
    how='left'
)
# Create third column = Final Approved LWD
merged_df[['EMP_STATUS', 'INACTIVEDATETIME','SYSTEM_DATE']] = merged_df[
    ['EMP_STATUS', 'INACTIVEDATETIME','SYSTEM_DATE']
].fillna('#N/A')
merged_df['INACTIVEDATETIME'] = pd.to_datetime(
    merged_df['INACTIVEDATETIME'], errors='coerce'
).dt.strftime('%d-%m-%Y')
merged_df['SYSTEM_DATE'] = pd.to_datetime(
    merged_df['SYSTEM_DATE'], errors='coerce'
).dt.strftime('%d-%m-%Y')
merged_df['SYSTEM_USER'] = merged_df['SYSTEM_USER'].fillna('')
merged_df['Final_Approved_LWD'] = merged_df['Final_Approved_LWD'].fillna('#N/A')
merged_df['Final_Approved_LWD'] = pd.to_datetime(
    merged_df['Final_Approved_LWD'], errors='coerce'
).dt.strftime('%d-%m-%Y')
merged_df['Date Of Resignation'] = pd.to_datetime(
    merged_df['Date Of Resignation'], errors='coerce'
).dt.strftime('%d-%m-%Y')
# Final Approved LWD logic
def compute_lwd(row):
    if row['EMP_STATUS'] == '#N/A':
        return '#N/A'
    elif row['EMP_STATUS'] == '75' or pd.isna(row['INACTIVEDATETIME']):
        return False
    elif  row['EMP_STATUS'] == '76' or row['INACTIVEDATETIME'] != row['Final_Approved_LWD']:
        return False
    else:
        return True

merged_df['LWD'] = merged_df.apply(compute_lwd, axis=1)

# Drop EMPNO (not needed in final output)
merged_df.drop(columns=['EMPNO'], inplace=True)

# Reorder columns → place new columns after Employee Code
cols = list(merged_df.columns)
emp_index = cols.index('Employee Code')
# Remove and reinsert in correct order
for col in ['EMP_STATUS', 'INACTIVEDATETIME', 'LWD']:
    cols.remove(col)
cols[emp_index+1:emp_index+1] = ['EMP_STATUS', 'INACTIVEDATETIME', 'LWD']
merged_df = merged_df[cols]
#print(merged_df)
sheet3_df = merged_df[
    (merged_df['EMP_STATUS'].astype(str) == '75') |
    (merged_df['LWD'].fillna(False).astype(bool) == False)
]
sheet3_df = sheet3_df.sort_values(
    by=['EMP_STATUS', 'LWD'],
    ascending=[False, False]
)

# Step 5: Write to Excel (Sheet2)
output_path = r"E:\Python_work\HR_SHEET\output.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    merged_df.to_excel(writer, sheet_name="Sheet1", index=False)  
    result_df.to_excel(writer, sheet_name="Sheet2", index=False)
    sheet3_df.to_excel(writer, sheet_name="Sheet3", index=False)

wb = load_workbook(output_path)
ws = wb["Sheet3"]

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

headers = [cell.value for cell in ws[1]]
emp_col = headers.index("EMP_STATUS") + 1
lwd_col = headers.index("LWD") + 1

for row in range(2, ws.max_row + 1):
    emp_status = str(ws.cell(row=row, column=emp_col).value)
    lwd = ws.cell(row=row, column=lwd_col).value

    # GREEN: EMP_STATUS = 75
    if emp_status == "75":
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = green_fill

    # ORANGE: EMP_STATUS = 76 AND LWD = False
    elif emp_status == "76" and lwd == False:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = orange_fill
    # 🔻 REMOVE COLUMNS FROM SHEET3
remove_cols = ["INACTIVEDATETIME", "LWD"]

# Get updated headers again
headers = [cell.value for cell in ws[1]]

# Find column indexes to delete (reverse order to avoid shifting issue)
cols_to_delete = [headers.index(col) + 1 for col in remove_cols if col in headers]
cols_to_delete.sort(reverse=True)

# Delete columns
for col_idx in cols_to_delete:
    ws.delete_cols(col_idx)


wb.save(output_path)
#print("✅ Done! Output saved with Sheet1, Sheet2 & Sheet3")

