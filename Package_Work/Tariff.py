import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
sys.path.append(r"E:\python")
from DbConfig.db_connectionlive import get_connection
from datetime import datetime
from database import result_df5
from combined import output_file
from Requiredparameter import hospital_id,tariffgroup_id #,service_center_id,department_id
import numpy as np
import warnings
warnings.filterwarnings("ignore")

# hospital_id = input("Enter  Tariff HOSPITAL_ID: ")
# tariffgroup_id = input("Tariffgroup_ID: ")
conn = get_connection()
if not conn:
    print("DB Connection Failed")
    exit()

query1 = f""" select name,id from orgstructure where isactive=1 and id='{hospital_id}' """
result_df1 = pd.read_sql(query1, conn)

query2 = f""" select tariffgroup_name,id from tariffgroup where id='{tariffgroup_id}' """
result_df2 = pd.read_sql(query2, conn)

# query3 = f""" SELECT id,prev_id,( id - prev_id ) AS cnt
# FROM(SELECT id,LAG(id, 1) OVER( ORDER BY id) AS prev_id 
# FROM tariff) list WHERE id - prev_id > 20000 FETCH FIRST 1 ROW ONLY """
# result_df3 = pd.read_sql(query3, conn)

conn.close()


# GET VALUES
# =========================

hospital_name = result_df1.iloc[0]['NAME']
tariffgroup_name = result_df2.iloc[0]['TARIFFGROUP_NAME']
result_df5['tariffid'] = result_df5['PREV_ID'] + 1
start_id = int(result_df5['tariffid'].iloc[0])

# Path--------------------

file_path = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"


df = pd.read_excel(file_path, sheet_name='rate')

# Room type columns
room_columns = ['OP','Day Care', 'Eco Celing', 'Twin Ceiling', 'Single Celing','Deluxe ceiling']

# Convert columns into rows
result = df.melt(
    id_vars=['SERVICE_CODE', 'SERVICE_NAME', 'SERVICE_MASTER_ID'],
    value_vars=room_columns,
    var_name='Room_Type',
    value_name='Amount'
)

# Remove blank/null values
result = result.dropna(subset=['Amount'])

# Optional: Remove zero values
result = result[result['Amount'] != 0]

# Final column order
result = result[[
    'SERVICE_CODE',
    'SERVICE_MASTER_ID',
    'SERVICE_NAME',
    'Room_Type',
    'Amount'
]]

#print(result)

# Save output
output_file = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    result.to_excel(writer, sheet_name='work_rate', index=False)

sheet1 = pd.read_excel(file_path, sheet_name='work_rate')

Service_id = sheet1['SERVICE_MASTER_ID']
SERVICE_NAME = sheet1['SERVICE_NAME']
Room_Type= sheet1['Room_Type'] 
Amount= sheet1['Amount'] 

room_name = np.select(
    [   Room_Type == 'Day Care',
        Room_Type == 'Single Celing',
        Room_Type == 'Twin Ceiling',
        Room_Type == 'Eco Celing',
        Room_Type == 'Deluxe ceiling'
    ],
    [
        'Day Care',
        'Single',
        'Twin Sharing',
        'Economy',
        'Deluxe'
    ],
    default=''
)

#Sheet Column parameter-----------------

sheet1['ID']  =range(start_id, start_id + len(sheet1))
sheet1['ORIGINALID']  =range(start_id, start_id + len(sheet1))
sheet1['TARIFFVERSION'] = '0'
sheet1['NAME'] = np.where(
    room_name != '',
    SERVICE_NAME.astype(str)
    + "/"
    + hospital_name
    + "/"
    + room_name
    + "/"
    + tariffgroup_name,

    SERVICE_NAME.astype(str)
    + "/"
    + hospital_name
    + "/"
    + tariffgroup_name
)
sheet1['TARIFFGROUP'] = 29385
sheet1['TOTALCHARGES'] = Amount
sheet1['ADVANCE'] = 0
sheet1['HOSPITALID'] = hospital_id
sheet1['CREATEDBY'] = 19413005
sheet1['UPDATEDBY'] = ''
sheet1['CREATEDDATETIME'] = datetime.today().strftime('%d-%m-%y')
sheet1['UPDATEDDATETIME'] = ''
sheet1['ACTIVE'] = 1
sheet1['TARIFFTYPE'] = 729645
sheet1['SERVICE_ID'] = Service_id
sheet1['TARIFFGROUPID'] = np.select(
    [ Room_Type == 'Day Care',
     Room_Type == 'Eco Celing',
    Room_Type ==  'Twin Ceiling',
    Room_Type == 'Single Celing',
    Room_Type == 'Deluxe ceiling'
        ],
    [ '1742381','1742374','1742375','1742377','1742378'],default='')
sheet1['TARFFIC_CLASS_TYPE'] = np.select(
    [ Room_Type == 'OP',],
    [ 'HOSPITAL'],default='CHARGECLASS')
sheet1['EFFECTIVEDATE'] = datetime.today().strftime('%d-%m-%y')
sheet1['ISFIXED'] = 'N'
sheet1['VALIDITY'] = '0'
sheet1['ADVANCETYPE'] = '0'
sheet1['TARFFIC_CLASS_ID'] = np.select(
    [ Room_Type == 'OP'
        ],
    [ hospital_id],default=tariffgroup_id)

sheet1['PERCENTAGE_TARIFF'] = ''
sheet1['CHARGE_PERHOUR'] = '0'
sheet1['DEPARTMENT_ID'] = ''
sheet1['ALIASCODE'] = ''
sheet1['ALIASNAME'] = ''
sheet1['NEWTARIFFGROUPID'] = ''
sheet1['SPONSOR_ID'] = ''

# Required column order
required_columns = [
'ID',
'ORIGINALID',
'TARIFFVERSION',
'NAME',
'TARIFFGROUP',
'TOTALCHARGES',
'ADVANCE',
'HOSPITALID',
'CREATEDBY',
'UPDATEDBY',
'CREATEDDATETIME',
'UPDATEDDATETIME',
'ACTIVE',
'TARIFFTYPE',
'SERVICE_ID',
'TARFFIC_CLASS_TYPE',
'TARFFIC_CLASS_ID',
'EFFECTIVEDATE',
'ISFIXED',
'VALIDITY',
'ADVANCETYPE',
'TARIFFGROUPID',
'PERCENTAGE_TARIFF',
'CHARGE_PERHOUR',
'DEPARTMENT_ID',
'ALIASCODE',
'ALIASNAME',
'NEWTARIFFGROUPID',
'SPONSOR_ID'
]

output_df = sheet1[required_columns]

# Export to Excel
#output_file = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    output_df.to_excel(writer, sheet_name='final_rate', index=False)


wb = load_workbook(file_path)

# Sheets to remove
sheets_to_remove = ['work_rate', 'rate']

# Remove sheets if they exist
for sheet in sheets_to_remove:
    if sheet in wb.sheetnames:
        del wb[sheet]

wb.save(file_path)