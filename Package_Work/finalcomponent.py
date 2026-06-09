import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
sys.path.append(r"E:\python")
from DbConfig.db_connectionlive import get_connection
from component import output_file
from database import result_df4
from Requiredparameter import tariffgroup_id
import numpy as np
import warnings
warnings.filterwarnings("ignore")

conn = get_connection()
if not conn:
    print("DB Connection Failed")
    exit()
file_path = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"
df=pd.read_excel(file_path, sheet_name='Comp3')
codes = df['Remarks'].dropna().astype(str).str.strip()

# Convert to SQL IN format → '1001','1002','1003'
sql_list = ",".join([f"'{c}'" for c in codes])

# Convert to SQL IN format → '1001','1002','1003'

# Step 3: Create Query
query = f""" select groupid,groupname from groups where ACTIVE=1 """

query2 = f""" select service_code,service_name,service_master_id from servicemaster where  is_active='Y' and service_code  in ({sql_list}) """
result_df = pd.read_sql(query, conn)
result_df2 = pd.read_sql(query2, conn)

conn.close()

result_df2['service_name'] = result_df2['SERVICE_NAME']
result_df2['service_master_id'] = result_df2['SERVICE_MASTER_ID']
result_df2['SERVICE_CODE'] = result_df2['SERVICE_CODE']

#result_df['groupid'] = result_df['groupid'].astype(str).str.strip()

group_map = dict(
    zip(
        result_df['GROUPNAME'].astype(str).str.strip(),
        result_df['GROUPID']
    )
)

name_map = dict(
    zip(
        result_df2['SERVICE_CODE'].astype(str).str.strip(),
        result_df2['SERVICE_NAME']
    )
)
componentid = dict(
    zip(
        result_df2['SERVICE_CODE'].astype(str).str.strip(),
        result_df2['SERVICE_MASTER_ID']
    )
)


# Read Excel file
sheet1 = pd.read_excel(file_path, sheet_name='Package')

sheet2 = pd.read_excel(file_path, sheet_name='Comp3')

#package_id = sheet1['PACKAGEID']
#Comp_name = sheet2['Component Name']
type = sheet2['Remarks']
daycare=sheet2['Day Care']
eco=sheet2['Eco Celing']
twin=sheet2['Twin Ceiling']
Single=sheet2['Single Celing'] 
Deluxe=sheet2['Deluxe ceiling']

# Convert numeric columns safely

cols = ['Day Care', 'Eco Celing', 'Twin Ceiling', 'Single Celing']
sheet2[cols] = (sheet2[cols].replace(['-', ' ', ''], 0) .fillna(0))

for col in cols:
    sheet2[col] = pd.to_numeric(sheet2[col], errors='coerce').fillna(0)

# REMOVE RECORDS
# Type = Qty AND Eco Celing = 0
# =========================
sheet2 = sheet2[
    ~(
        (sheet2['Type'].astype(str).str.strip() == 'Qty') &
        (sheet2['Eco Celing'] <= 0)
    )
].copy()

Package_id = dict(
    zip(sheet1['PACKAGECODE'],
        sheet1['PACKAGEID'])
)

# result_df4['Componentid'] = result_df4['PREV_ID'] + 1
# start_id = int(result_df4['Componentid'].iloc[0])

sheet2['ID'] =''
sheet2['COMPONENTCODE'] = ''
sheet2['COMPONENTTYPE'] = np.select(
    [ sheet2['Component Name'] == 'IMPLANT 40',
     sheet2['Component Name'] == 'DrugsAndMaterials',
        ],
    [ '1052375','1052379'],default='1052373')
sheet2['COMPONENTNAME'] = sheet2['Remarks'].astype(str).str.strip().map(name_map)
sheet2['COMPONENTGROUPID'] = sheet2['Component Name'].astype(str).str.strip().map(group_map)
sheet2['COMPONENTID'] = sheet2['Remarks'].astype(str).str.strip().map(componentid)
sheet2['ISAUTOALLOCATION'] = '0'
sheet2['AMOUNTLIMIT'] = ''
sheet2['QTYLIMIT'] = np.where(
    (sheet2['Type']== 'Qty') &
    (sheet2['Eco Celing']> 0) &
    (sheet2['Twin Ceiling'] > 0) &
    (sheet2['Single Celing']> 0),
    
    sheet2['Eco Celing'].astype(float).astype(int),
    
    '')
sheet2['ISEXCLUDED'] = np.select(
    [ sheet2['Component Name'] == 'IMPLANT 40' ],
    [ '1'],default='')
sheet2['LIST_INDEX'] = ''
sheet2['PACKAGEID'] = sheet2['PKG code'].map(Package_id)
sheet2['EXEMPTION'] = '0'
sheet2['ACTIVE'] = '1'
sheet2['ISPACKAGECOMPONENT'] = '0'
sheet2['TARIFFCLASSTYPE'] = ''
sheet2['TARIFFCLASSVALUE'] = ''
sheet2['GENERIC'] = '1'
sheet2['COMPONENTPRICELIMIT'] = ''
sheet2['CPT_CODE'] = ''
sheet2['TARIFF_GROUP_ID'] = ''
sheet2['PERCENTAGE'] = '0'
sheet2['COMPONENTDISCOUNTEDPRICE'] = ''

required_columns = [
    'ID',
'COMPONENTCODE',
'COMPONENTTYPE',
'COMPONENTNAME',
'COMPONENTGROUPID',
'COMPONENTID',
'ISAUTOALLOCATION',
'AMOUNTLIMIT',
'QTYLIMIT',
'ISEXCLUDED',
'LIST_INDEX',
'PACKAGEID',
'EXEMPTION',
'ACTIVE',
'ISPACKAGECOMPONENT',
'TARIFFCLASSTYPE',
'TARIFFCLASSVALUE',
'GENERIC',
'COMPONENTPRICELIMIT',
'CPT_CODE',
'TARIFF_GROUP_ID',
'PERCENTAGE',
'COMPONENTDISCOUNTEDPRICE'
]

output_df = sheet2[required_columns]

# Export to Excel
output_file = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    output_df.to_excel(writer, sheet_name='Comp4', index=False)

 #Comp5----------------------------------------------------------------------------------   
sheet3= pd.read_excel(file_path, sheet_name='Packagew_rate')

Room_Type= sheet3['Room_Type'] 
#TARIFF_GROUP_ID = input("Enter TARIFF_GROUP_ID: ")

sheet3['ID'] =''
sheet3['COMPONENTCODE'] = ''
sheet3['COMPONENTTYPE'] = np.select(
    [ sheet3['COMPONENTTYPE'] == 'IMPLANT 40',
     sheet3['COMPONENTTYPE'] == 'DrugsAndMaterials',
        ],
    [ '1052375','1052379'],default='1052373')
sheet3['COMPONENTNAME'] =sheet3['Component Name']
sheet3['COMPONENTGROUPID'] = sheet3['Component Name'].astype(str).str.strip().map(group_map)
sheet3['COMPONENTID'] = '0'
sheet3['ISAUTOALLOCATION'] = '0'
sheet3['AMOUNTLIMIT'] = sheet3['Amount']
sheet3['QTYLIMIT'] = '0'
sheet3['ISEXCLUDED'] = ''
sheet3['LIST_INDEX'] = ''
sheet3['PACKAGEID'] = sheet3['PKG code'].map(Package_id)
sheet3['EXEMPTION'] = '0'
sheet3['ACTIVE'] = '1'
sheet3['ISPACKAGECOMPONENT'] = '0'
sheet3['TARIFFCLASSTYPE'] = 'SPONSOR_TARIFF_GROUP'
sheet3['TARIFFCLASSVALUE'] = tariffgroup_id
sheet3['GENERIC'] = '1'
sheet3['COMPONENTPRICELIMIT'] = ''
sheet3['CPT_CODE'] = ''
sheet3['TARIFF_GROUP_ID'] = np.select(
    [ Room_Type == 'Day Care',
     Room_Type == 'Eco Celing',
    Room_Type ==  'Twin Ceiling',
    Room_Type == 'Single Celing' ,
    Room_Type == 'Deluxe ceiling'    
        ],
    [ '1742381','1742374','1742375','1742377','1742378'],default='')
sheet3['PERCENTAGE'] = '0'
sheet3['COMPONENTDISCOUNTEDPRICE'] = ''

required_columns = [
    'ID',
'COMPONENTCODE',
'COMPONENTTYPE',
'COMPONENTNAME',
'COMPONENTGROUPID',
'COMPONENTID',
'ISAUTOALLOCATION',
'AMOUNTLIMIT',
'QTYLIMIT',
'ISEXCLUDED',
'LIST_INDEX',
'PACKAGEID',
'EXEMPTION',
'ACTIVE',
'ISPACKAGECOMPONENT',
'TARIFFCLASSTYPE',
'TARIFFCLASSVALUE',
'GENERIC',
'COMPONENTPRICELIMIT',
'CPT_CODE',
'TARIFF_GROUP_ID',
'PERCENTAGE',
'COMPONENTDISCOUNTEDPRICE'
]

output_df = sheet3[required_columns]

# Export to Excel
with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    output_df.to_excel(writer, sheet_name='Comp5', index=False)


sheet7 = pd.read_excel(output_file, sheet_name="Comp1")
sheet4 = pd.read_excel(output_file, sheet_name="Comp2")
sheet3 = pd.read_excel(output_file, sheet_name="Comp4")
sheet6 = pd.read_excel(output_file, sheet_name="Comp5")



combined_df = pd.concat([sheet7, sheet4,sheet3, sheet6], ignore_index=True)
result_df4['ID'] = result_df4['PREV_ID'] + 1
start_id = int(result_df4['ID'].iloc[0])

# Generate unique IDs
combined_df['ID'] = range(
    start_id,
    start_id + len(combined_df)
)

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    combined_df.to_excel(writer, sheet_name='finalcompnent', index=False)


wb = load_workbook(file_path)

# Sheets to remove
sheets_to_remove = ['Comp1', 'Comp2', 'Comp3','Packagew_rate','Comp4', 'Comp5','Working']

# Remove sheets if they exist
for sheet in sheets_to_remove:
    if sheet in wb.sheetnames:
        del wb[sheet]

wb.save(file_path)