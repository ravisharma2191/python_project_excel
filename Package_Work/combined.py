import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from database import result_df2,result_df3
from datetime import datetime
from Working import output_file
from Surgery import output_file as filnalsheet
from Requiredparameter import hospital_id,tariffgroup_id,service_center_id,department_id
import warnings
warnings.filterwarnings("ignore")

# Excel file path
file_path = r"C:\Users\software.support\Desktop\New folder\Output-Package.xlsx"

# Read both worksheets
sheet1 = pd.read_excel(file_path, sheet_name="Service")
sheet2 = pd.read_excel(file_path, sheet_name="Surgery")

sheet3 = pd.read_excel(file_path, sheet_name="packagerate")
sheet4 = pd.read_excel(file_path, sheet_name="servicerate")

# Combine both sheets
combined_df = pd.concat([sheet1, sheet2], ignore_index=True)
combined_df1 = pd.concat([sheet3, sheet4], ignore_index=True)

result_df2['service_master_id'] = result_df2['PREV_ID'] + 1
start_id = int(result_df2['service_master_id'].iloc[0])

# Generate unique IDs
combined_df['SERVICE_MASTER_ID'] = range(
    start_id,
    start_id + len(combined_df)
)

# Save combined data to new Excel file
output_file = filnalsheet

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    combined_df.to_excel(writer, sheet_name='Servicemaster', index=False)


    output_file = filnalsheet

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    combined_df1.to_excel(writer, sheet_name='rate', index=False)

# Read sheets
package_df = pd.read_excel(file_path, sheet_name="Package")
service_df = pd.read_excel(file_path, sheet_name="Servicemaster")
rate_df = pd.read_excel(file_path, sheet_name="rate")


# service masterid in package sheet.----------------------------------------
# Create mapping
service_map = dict(
    zip(service_df['SERVICE_CODE'],
        service_df['SERVICE_MASTER_ID'])
)

# Update existing column
package_df['SERVICEMASTERID'] = package_df['PACKAGECODE'].map(service_map)

# Save output
#package_df.to_excel("Output-Package.xlsx", index=False)

output_file = filnalsheet

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    package_df.to_excel(writer, sheet_name='Package', index=False)

#service master id in rate sheet-------------------------------------------------
merged_df = rate_df.merge(
    service_df[['SERVICE_MASTER_ID', 'SERVICE_CODE']],
    left_on='SERVICE_CODE',
    right_on='SERVICE_CODE',
    how='left'
)

# Rename column if needed
merged_df.rename(columns={
    'SERVICE_MASTER_ID': 'SERVICE_MASTER_ID'
}, inplace=True)

# Optional: remove extra column
#merged_df.drop(columns=['SERVICE_CODE'], inplace=True)

output_file = filnalsheet

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    merged_df.to_excel(writer, sheet_name='rate', index=False)


#Location Sheet generate-------------------------------------------

result_df3['service_loc_id'] = result_df3['PREV_ID'] + 1
start_id1 = int(result_df3['service_loc_id'].iloc[0])

# hospital_id = input("Enter Location HOSPITAL_ID: ")
# department_id = input("Enter DEPARTMENT_ID: ")
# service_center_id = input("Enter SERVICE_CENTER_ID: ")

sheet6 = pd.read_excel(file_path, sheet_name='Servicemaster')
SERVICE_ID = sheet6['SERVICE_MASTER_ID']

sheet6['SERVICE_LOC_ID'] =range(start_id1, start_id1 + len(sheet6))
sheet6['HOSPITAL_ID'] =hospital_id
sheet6['DEPARTMENT_ID'] =department_id
sheet6['SERVICE_CENTER_ID'] =service_center_id
sheet6['IS_PRIMARY_LOCATION'] ='1'
sheet6['OP_BILL_TYPE'] ='0'
sheet6['IP_BILL_TYPE'] ='0'
sheet6['SERVICE_ID'] =SERVICE_ID
sheet6['TARIFFID'] =''
sheet6['CREATEDBY'] =19413005
sheet6['CREATEDDT'] =datetime.today().strftime('%d-%m-%y')
sheet6['UPDATEDBY'] =''
sheet6['UPDATEDDT'] =''
sheet6['ISSUSPEND'] ='0'
sheet6['SUSPENDDATE'] =''
sheet6['ECONOMICCLASS'] =''

# Required column order
required_columns = [
'SERVICE_LOC_ID',
'HOSPITAL_ID',
'DEPARTMENT_ID',
'SERVICE_CENTER_ID',
'IS_PRIMARY_LOCATION',
'OP_BILL_TYPE',
'IP_BILL_TYPE',
'SERVICE_ID',
'TARIFFID',
'CREATEDBY',
'CREATEDDT',
'UPDATEDBY',
'UPDATEDDT',
'ISSUSPEND',
'SUSPENDDATE',
'ECONOMICCLASS'
]
output_df = sheet6[required_columns]

# Export to Excel
output_file = filnalsheet

with pd.ExcelWriter(output_file,engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    output_df.to_excel(writer, sheet_name='locationmaster', index=False)



 #Remove working Sheet from output file---------------------------
#Load workbook
wb = load_workbook(file_path)

# Sheets to remove
sheets_to_remove = ['Service', 'Surgery', 'packagerate', 'servicerate']

# Remove sheets if they exist
for sheet in sheets_to_remove:
    if sheet in wb.sheetnames:
        del wb[sheet]

wb.save(file_path)